import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_aula05_workbook():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles & Colors (Navy & Sky Blue aeronautical theme)
    font_family = "Segoe UI"
    
    NAVY_HEX = "0F2043"
    STEEL_BLUE_HEX = "0072CE"
    ICE_BLUE_HEX = "EBF3FA"
    LIGHT_GRAY_HEX = "F8FAFC"
    BORDER_GRAY_HEX = "CBD5E1"
    ACCENT_GREEN_HEX = "107C41"
    ALERT_YELLOW_HEX = "FFF3CD"
    ALERT_RED_HEX = "FEE2E2"
    
    font_title = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=11, bold=True, color=NAVY_HEX)
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_header_sec = Font(name=font_family, size=10, bold=True, color=NAVY_HEX)
    font_data = Font(name=font_family, size=10, color="1E293B")
    font_bold = Font(name=font_family, size=10, bold=True, color="1E293B")
    font_formula = Font(name="Consolas", size=9, bold=True, color="004085")
    font_note = Font(name=font_family, size=9, italic=True, color="64748B")
    
    fill_title = PatternFill(start_color=NAVY_HEX, end_color=NAVY_HEX, fill_type="solid")
    fill_header = PatternFill(start_color=NAVY_HEX, end_color=NAVY_HEX, fill_type="solid")
    fill_header_accent = PatternFill(start_color=STEEL_BLUE_HEX, end_color=STEEL_BLUE_HEX, fill_type="solid")
    fill_zebra = PatternFill(start_color=LIGHT_GRAY_HEX, end_color=LIGHT_GRAY_HEX, fill_type="solid")
    fill_formula_highlight = PatternFill(start_color=ICE_BLUE_HEX, end_color=ICE_BLUE_HEX, fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_card = PatternFill(start_color=ICE_BLUE_HEX, end_color=ICE_BLUE_HEX, fill_type="solid")
    fill_alert = PatternFill(start_color=ALERT_YELLOW_HEX, end_color=ALERT_YELLOW_HEX, fill_type="solid")
    fill_outlier = PatternFill(start_color=ALERT_RED_HEX, end_color=ALERT_RED_HEX, fill_type="solid")

    thin_border_side = Side(border_style="thin", color=BORDER_GRAY_HEX)
    double_border_side = Side(border_style="double", color="1E293B")
    thick_bottom_side = Side(border_style="medium", color=NAVY_HEX)
    
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=double_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CURRENCY_FORMAT = 'R$ #,##0.00'
    NUMBER_FORMAT = '#,##0'

    # Fleet Raw Data
    frota_dados = [
        ("PR-ABC", "Cessna 172", 4, 150.00, 120, 11.50),
        ("PT-KRT", "Piper Seneca", 10, 280.00, 350, 11.50),
        ("PR-FTE", "Beech Baron", 7, 320.00, 280, 11.50),
        ("PP-ZUL", "Cirrus SR22", 3, 220.00, 160, 11.50),
        ("PR-MNT", "Cessna 152", 15, 120.00, 90, 11.50),
    ]

    # Training Raw Data
    treinamento_dados = [
        ("Lucas Silveira", 18.5, 34.0),
        ("Mariana Duarte", 25.0, 36.5),
        ("Carlos Eduardo", 12.0, 32.0),
        ("Beatriz Ramos", 30.5, 38.0),
        ("Rodrigo Nogueira", 22.0, 35.0),
        ("Gabriel Souza", 15.5, 33.5),
    ]

    def autofit(ws, min_col=1, max_col=None, padding=3):
        if max_col is None:
            max_col = ws.max_column
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if row == 1 and ws.cell(row=row, column=col).coordinate != 'A1':
                    continue
                if row == 1 and col == 1:
                    continue
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    val_str = "R$ 120.000,00"
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + padding, 12)

    # =========================================================================
    # ABA 0: INSTRUÇÕES E ATALHOS
    # =========================================================================
    ws0 = wb.create_sheet(title="00_Instrucoes_e_Atalhos")
    ws0.views.sheetView[0].showGridLines = True
    
    ws0.merge_cells("A1:G1")
    ws0["A1"] = "INF-117 INFORMÁTICA APLICADA A AERONÁUTICA — AULA 05: GUIA DE ATIVIDADES NO EXCEL"
    ws0["A1"].font = font_title; ws0["A1"].fill = fill_title; ws0["A1"].alignment = align_center
    ws0.row_dimensions[1].height = 32

    ws0["A3"] = "Roteiro Prático Passo a Passo — Navegação entre as Abas da Planilha"
    ws0["A3"].font = font_section

    instrucoes = [
        ("Aba 1: Ativ1_Estrutura_Hangar", "Criação do relatório do Aeroclube de Sorocaba, definição do título e dos cabeçalhos das 9 colunas."),
        ("Aba 2: Ativ2_Dados_Iniciais", "Inserção dos dados operacionais da frota de aeronaves (Prefixos, Modelos, Diárias, Preços e Combustível)."),
        ("Aba 3: Ativ3_Formulas_Operadores", "Aplicação de operadores aritméticos básicos (* e +) para cálculo de Subtotais e Custo Total."),
        ("Aba 4: Ativ4_Formatacao_e_Totais", "Aplicação de formatação de moeda oficial (R$), alinhamentos, bordas e cálculo do TOTAL GERAL via =SOMA()."),
        ("Aba 5: Precedencia_Potencia_e_ISA", "Exploração aprofundada dos operadores ^ e -: Pressão Dinâmica (0.5*rho*V^2) e Gradiente ISA (15 - 0.0065*h)."),
        ("Aba 6: Media_vs_Mediana_OS", "Estudo de caso aeronáutico: Análise de ordens de serviço (OS) e o impacto de outliers em =MÉDIA() vs =MED()."),
        ("Aba 7: Exercicio_Fixacao_Consumo", "Construção da planilha de consumo em horas de voo de instrução com =SOMA() e =MÉDIA()."),
        ("Aba 8: Gabarito_Completo", "Painel consolidado completo com todas as fórmulas ativas e comparativos práticos."),
    ]

    ws0.cell(row=5, column=1, value="Aba / Etapa").font = font_header
    ws0.cell(row=5, column=1).fill = fill_header_accent
    ws0.cell(row=5, column=1).alignment = align_center
    ws0.cell(row=5, column=1).border = border_cell

    ws0.merge_cells("B5:G5")
    ws0.cell(row=5, column=2, value="Descrição da Atividade & Objetivos de Aprendizagem").font = font_header
    ws0.cell(row=5, column=2).fill = fill_header_accent
    ws0.cell(row=5, column=2).alignment = align_left
    ws0.cell(row=5, column=2).border = border_cell

    for i, (aba_nome, desc) in enumerate(instrucoes, start=6):
        c1 = ws0.cell(row=i, column=1, value=aba_nome)
        c1.font = font_bold; c1.fill = fill_zebra if i % 2 == 0 else PatternFill(fill_type=None)
        c1.alignment = align_left; c1.border = border_cell

        ws0.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        c2 = ws0.cell(row=i, column=2, value=desc)
        c2.font = font_data; c2.fill = fill_zebra if i % 2 == 0 else PatternFill(fill_type=None)
        c2.alignment = align_left; c2.border = border_cell
        ws0.row_dimensions[i].height = 22

    # Atalhos
    ws0["A16"] = "Tabela de Atalhos Rápidos Essenciais no MS Excel"
    ws0["A16"].font = font_section

    atalhos = [
        ("Ctrl + Z", "Desfazer a última ação"),
        ("Ctrl + Y", "Refazer a última ação"),
        ("Alt + =", "Insere automaticamente a fórmula de Soma (=SOMA)"),
        ("Ctrl + Shift + $", "Aplica a formatação de Moeda (R$) nas células selecionadas"),
        ("Ctrl + Shift + %", "Aplica a formatação de Porcentagem (%)"),
        ("=MED(intervalo)", "Calcula a Mediana (valor central, imune a distorções por outliers)"),
        ("=MÉDIA(intervalo)", "Calcula a Média Aritmética (soma dividida pela quantidade)"),
        ("Duplo clique na borda", "Autoajusta a largura da coluna ao conteúdo mais extenso"),
    ]

    ws0.cell(row=18, column=1, value="Atalho / Sintaxe").font = font_header
    ws0.cell(row=18, column=1).fill = fill_header
    ws0.cell(row=18, column=1).alignment = align_center
    ws0.cell(row=18, column=1).border = border_cell

    ws0.merge_cells("B18:D18")
    ws0.cell(row=18, column=2, value="Função no Excel").font = font_header
    ws0.cell(row=18, column=2).fill = fill_header
    ws0.cell(row=18, column=2).alignment = align_left
    ws0.cell(row=18, column=2).border = border_cell

    for i, (tecla, func) in enumerate(atalhos, start=19):
        c1 = ws0.cell(row=i, column=1, value=tecla)
        c1.font = font_formula; c1.fill = fill_formula_highlight; c1.alignment = align_center; c1.border = border_cell

        ws0.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        c2 = ws0.cell(row=i, column=2, value=func)
        c2.font = font_data; c2.alignment = align_left; c2.border = border_cell
        ws0.row_dimensions[i].height = 20

    ws0.column_dimensions["A"].width = 34
    ws0.column_dimensions["B"].width = 25
    ws0.column_dimensions["C"].width = 25
    ws0.column_dimensions["D"].width = 25

    # =========================================================================
    # ABA 1: ATIVIDADE 1 — ESTRUTURA DO RELATÓRIO DO HANGAR
    # =========================================================================
    ws1 = wb.create_sheet(title="Ativ1_Estrutura_Hangar")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:I1")
    ws1["A1"] = "AEROCLUBE DE SOROCABA — CONTROLE DE HANGARAGEM E ABASTECIMENTO"
    ws1["A1"].font = font_title; ws1["A1"].fill = fill_title; ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:I2")
    ws1["A2"] = "Etapa 1: Definição do Título da Planilha e Estrutura das 9 Colunas de Controle Operacional"
    ws1["A2"].font = font_note; ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 18

    headers_hangar = [
        "Prefixo", "Modelo da Aeronave", "Diárias Hangar", "Valor da Diária (R$)",
        "Litros Abastecidos (L)", "Preço por Litro (R$)", "Subtotal Hangar (R$)",
        "Subtotal Combustível (R$)", "Custo Total (R$)"
    ]

    for col_idx, text in enumerate(headers_hangar, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=text)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_header; cell.border = border_cell
    ws1.row_dimensions[3].height = 28

    for r in range(4, 9):
        for c in range(1, 10):
            ws1.cell(row=r, column=c).border = border_cell
        ws1.row_dimensions[r].height = 20

    ws1.merge_cells("A11:I13")
    ws1["A11"] = "ORIENTAÇÃO DA ATIVIDADE 1:\n• Nesta etapa inicial, definimos o cabeçalho principal na linha 1 e as 9 colunas na linha 3.\n• Observe que as colunas A a F receberão dados brutos de entrada, enquanto as colunas G, H e I serão campos calculados."
    ws1["A11"].font = font_section; ws1["A11"].fill = fill_card
    ws1["A11"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    autofit(ws1, 1, 9)

    # =========================================================================
    # ABA 2: ATIVIDADE 2 — DADOS INICIAIS
    # =========================================================================
    ws2 = wb.create_sheet(title="Ativ2_Dados_Iniciais")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:I1")
    ws2["A1"] = "AEROCLUBE DE SOROCABA — CONTROLE DE HANGARAGEM E ABASTECIMENTO"
    ws2["A1"].font = font_title; ws2["A1"].fill = fill_title; ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 30

    for col_idx, text in enumerate(headers_hangar, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_header; cell.border = border_cell
    ws2.row_dimensions[3].height = 28

    for row_idx, row_data in enumerate(frota_dados, start=4):
        ws2.cell(row=row_idx, column=1, value=row_data[0]).font = font_data; ws2.cell(row=row_idx, column=1).alignment = align_center; ws2.cell(row=row_idx, column=1).border = border_cell
        ws2.cell(row=row_idx, column=2, value=row_data[1]).font = font_data; ws2.cell(row=row_idx, column=2).alignment = align_left; ws2.cell(row=row_idx, column=2).border = border_cell
        ws2.cell(row=row_idx, column=3, value=row_data[2]).font = font_data; ws2.cell(row=row_idx, column=3).alignment = align_center; ws2.cell(row=row_idx, column=3).border = border_cell
        ws2.cell(row=row_idx, column=4, value=row_data[3]).font = font_data; ws2.cell(row=row_idx, column=4).alignment = align_right; ws2.cell(row=row_idx, column=4).border = border_cell
        ws2.cell(row=row_idx, column=5, value=row_data[4]).font = font_data; ws2.cell(row=row_idx, column=5).alignment = align_center; ws2.cell(row=row_idx, column=5).border = border_cell
        ws2.cell(row=row_idx, column=6, value=row_data[5]).font = font_data; ws2.cell(row=row_idx, column=6).alignment = align_right; ws2.cell(row=row_idx, column=6).border = border_cell
        for col_c in range(7, 10):
            ws2.cell(row=row_idx, column=col_c).border = border_cell
        ws2.row_dimensions[row_idx].height = 20

    ws2.merge_cells("A11:I13")
    ws2["A11"] = "ORIENTAÇÃO DA ATIVIDADE 2:\n• Inserimos os dados das 5 aeronaves operadas na semana (linhas 4 a 8).\n• Note que os valores monetários ainda estão como números puros (ex: 150 e 11.5). Na próxima etapa iremos criar as fórmulas nas colunas G, H e I."
    ws2["A11"].font = font_section; ws2["A11"].fill = fill_card
    ws2["A11"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    autofit(ws2, 1, 9)

    # =========================================================================
    # ABA 3: ATIVIDADE 3 — FÓRMULAS ARITMÉTICAS
    # =========================================================================
    ws3 = wb.create_sheet(title="Ativ3_Formulas_Operadores")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:I1")
    ws3["A1"] = "AEROCLUBE DE SOROCABA — CONTROLE DE HANGARAGEM E ABASTECIMENTO"
    ws3["A1"].font = font_title; ws3["A1"].fill = fill_title; ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 30

    for col_idx, text in enumerate(headers_hangar, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=text)
        cell.font = font_header; cell.fill = fill_header_accent if col_idx >= 7 else fill_header; cell.alignment = align_header; cell.border = border_cell
    ws3.row_dimensions[3].height = 28

    for row_idx, row_data in enumerate(frota_dados, start=4):
        ws3.cell(row=row_idx, column=1, value=row_data[0]).alignment = align_center
        ws3.cell(row=row_idx, column=2, value=row_data[1]).alignment = align_left
        ws3.cell(row=row_idx, column=3, value=row_data[2]).alignment = align_center
        ws3.cell(row=row_idx, column=4, value=row_data[3]).alignment = align_right
        ws3.cell(row=row_idx, column=5, value=row_data[4]).alignment = align_center
        ws3.cell(row=row_idx, column=6, value=row_data[5]).alignment = align_right
        
        cG = ws3.cell(row=row_idx, column=7, value=f"=C{row_idx}*D{row_idx}")
        cH = ws3.cell(row=row_idx, column=8, value=f"=E{row_idx}*F{row_idx}")
        cI = ws3.cell(row=row_idx, column=9, value=f"=G{row_idx}+H{row_idx}")
        
        cG.fill = fill_formula_highlight; cG.font = font_bold; cG.alignment = align_right
        cH.fill = fill_formula_highlight; cH.font = font_bold; cH.alignment = align_right
        cI.fill = fill_formula_highlight; cI.font = font_bold; cI.alignment = align_right

        for c_idx in range(1, 10):
            cell = ws3.cell(row=row_idx, column=c_idx)
            cell.border = border_cell
            if c_idx < 7:
                cell.font = font_data
        ws3.row_dimensions[row_idx].height = 20

    ws3.merge_cells("A11:I14")
    ws3["A11"] = "FÓRMULAS CONSTRUÍDAS NESTA ETAPA:\n• Subtotal Hangar (G4): =C4*D4  (Diárias × Valor da Diária)\n• Subtotal Combustível (H4): =E4*F4  (Litros × Preço por Litro)\n• Custo Total (I4): =G4+H4  (Subtotal Hangar + Subtotal Combustível)\n• Alça de Preenchimento: Selecione G4:I4 e dê duplo clique no quadradinho inferior direito para replicar até a linha 8!"
    ws3["A11"].font = font_section; ws3["A11"].fill = fill_card
    ws3["A11"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    autofit(ws3, 1, 9)

    # =========================================================================
    # ABA 4: ATIVIDADE 4 — FORMATAÇÃO PROFISSIONAL E TOTAIS
    # =========================================================================
    ws4 = wb.create_sheet(title="Ativ4_Formatacao_e_Totais")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:I1")
    ws4["A1"] = "AEROCLUBE DE SOROCABA — CONTROLE DE HANGARAGEM E ABASTECIMENTO"
    ws4["A1"].font = font_title; ws4["A1"].fill = fill_title; ws4["A1"].alignment = align_center
    ws4.row_dimensions[1].height = 30

    for col_idx, text in enumerate(headers_hangar, start=1):
        cell = ws4.cell(row=3, column=col_idx, value=text)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_header; cell.border = border_cell
    ws4.row_dimensions[3].height = 28

    for row_idx, row_data in enumerate(frota_dados, start=4):
        zebra = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        c = ws4.cell(row=row_idx, column=1, value=row_data[0]); c.font = font_bold; c.alignment = align_center; c.border = border_cell; c.fill = zebra
        c = ws4.cell(row=row_idx, column=2, value=row_data[1]); c.font = font_data; c.alignment = align_left; c.border = border_cell; c.fill = zebra
        c = ws4.cell(row=row_idx, column=3, value=row_data[2]); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = NUMBER_FORMAT
        c = ws4.cell(row=row_idx, column=4, value=row_data[3]); c.font = font_data; c.alignment = align_right; c.border = border_cell; c.fill = zebra; c.number_format = CURRENCY_FORMAT
        c = ws4.cell(row=row_idx, column=5, value=row_data[4]); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = NUMBER_FORMAT
        c = ws4.cell(row=row_idx, column=6, value=row_data[5]); c.font = font_data; c.alignment = align_right; c.border = border_cell; c.fill = zebra; c.number_format = CURRENCY_FORMAT
        
        cG = ws4.cell(row=row_idx, column=7, value=f"=C{row_idx}*D{row_idx}"); cG.font = font_data; cG.alignment = align_right; cG.border = border_cell; cG.fill = zebra; cG.number_format = CURRENCY_FORMAT
        cH = ws4.cell(row=row_idx, column=8, value=f"=E{row_idx}*F{row_idx}"); cH.font = font_data; cH.alignment = align_right; cH.border = border_cell; cH.fill = zebra; cH.number_format = CURRENCY_FORMAT
        cI = ws4.cell(row=row_idx, column=9, value=f"=G{row_idx}+H{row_idx}"); cI.font = font_bold; cI.alignment = align_right; cI.border = border_cell; cI.fill = zebra; cI.number_format = CURRENCY_FORMAT
        ws4.row_dimensions[row_idx].height = 20

    # Totais
    ws4.merge_cells("A9:F9")
    ws4["A9"] = "TOTAL GERAL"; ws4["A9"].font = font_bold; ws4["A9"].alignment = align_right; ws4["A9"].fill = fill_total
    for col_c in range(1, 7):
        ws4.cell(row=9, column=col_c).border = border_total; ws4.cell(row=9, column=col_c).fill = fill_total

    ws4["G9"] = "=SUM(G4:G8)"; ws4["G9"].font = font_bold; ws4["G9"].alignment = align_right; ws4["G9"].border = border_total; ws4["G9"].fill = fill_total; ws4["G9"].number_format = CURRENCY_FORMAT
    ws4["H9"] = "=SUM(H4:H8)"; ws4["H9"].font = font_bold; ws4["H9"].alignment = align_right; ws4["H9"].border = border_total; ws4["H9"].fill = fill_total; ws4["H9"].number_format = CURRENCY_FORMAT
    ws4["I9"] = "=SUM(I4:I8)"; ws4["I9"].font = Font(name=font_family, size=10, bold=True, color=NAVY_HEX); ws4["I9"].alignment = align_right; ws4["I9"].border = border_total; ws4["I9"].fill = fill_total; ws4["I9"].number_format = CURRENCY_FORMAT
    ws4.row_dimensions[9].height = 22

    ws4.merge_cells("A11:I13")
    ws4["A11"] = "RESULTADO FINAL DA ATIVIDADE 4:\n• Formatação em Moeda Oficial (R$ #.##0,00) aplicada a todas as colunas de valor.\n• Linha 9 consolidada com a função =SOMA() em G9, H9 e I9.\n• Bordas duplas na parte inferior indicam o encerramento do balanço financeiro."
    ws4["A11"].font = font_section; ws4["A11"].fill = fill_card
    ws4["A11"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    autofit(ws4, 1, 9)

    # =========================================================================
    # ABA 5: EXPLORAÇÃO DE PRECEDÊNCIA, POTÊNCIA (^) E SUBTRAÇÃO (-) NA AERONÁUTICA
    # =========================================================================
    ws_prec = wb.create_sheet(title="Precedencia_Potencia_e_ISA")
    ws_prec.views.sheetView[0].showGridLines = True

    ws_prec.merge_cells("A1:G1")
    ws_prec["A1"] = "LABORATÓRIO DE CÁLCULO: PRECEDÊNCIA MATEMÁTICA, OPERADORES ^ E - EM FÍSICA AERONÁUTICA"
    ws_prec["A1"].font = font_title; ws_prec["A1"].fill = fill_title; ws_prec["A1"].alignment = align_center
    ws_prec.row_dimensions[1].height = 30

    # Tabela 1: Pressão Dinâmica q = 0.5 * rho * V^2
    ws_prec["A3"] = "1. Cálculo de Pressão Dinâmica Aerodinâmica: q = 0.5 * ρ * V²"
    ws_prec["A3"].font = font_section

    headers_q = ["Aeronave", "Densidade do Ar ρ (kg/m³)", "Velocidade V (m/s)", "Fórmula no Excel", "Pressão Dinâmica q (Pa)"]
    for col_idx, text in enumerate(headers_q, start=1):
        c = ws_prec.cell(row=4, column=col_idx, value=text)
        c.font = font_header; c.fill = fill_header; c.alignment = align_header; c.border = border_cell
    ws_prec.row_dimensions[4].height = 25

    aero_q_dados = [
        ("Cessna 152 (Cruzeiro)", 1.225, 51.4, "=0.5*B5*C5^2"),
        ("Cessna 172 (Cruzeiro)", 1.225, 61.7, "=0.5*B6*C6^2"),
        ("Cirrus SR22 (Cruzeiro)", 1.100, 92.6, "=0.5*B7*C7^2"),
        ("Embraer E195-E2 (Cruzeiro FL350)", 0.380, 240.0, "=0.5*B8*C8^2"),
    ]

    for row_idx, (aviao, rho, v, _) in enumerate(aero_q_dados, start=5):
        zebra = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
        ws_prec.cell(row=row_idx, column=1, value=aviao).font = font_bold; ws_prec.cell(row=row_idx, column=1).alignment = align_left; ws_prec.cell(row=row_idx, column=1).border = border_cell; ws_prec.cell(row=row_idx, column=1).fill = zebra
        c = ws_prec.cell(row=row_idx, column=2, value=rho); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.000'
        c = ws_prec.cell(row=row_idx, column=3, value=v); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.0'
        c = ws_prec.cell(row=row_idx, column=4, value=f"=0.5 * B{row_idx} * C{row_idx}^2"); c.font = font_formula; c.alignment = align_center; c.border = border_cell; c.fill = fill_formula_highlight
        c = ws_prec.cell(row=row_idx, column=5, value=f"=0.5*B{row_idx}*C{row_idx}^2"); c.font = font_bold; c.alignment = align_right; c.border = border_cell; c.fill = zebra; c.number_format = '#,##0.0 "Pa"'
        ws_prec.row_dimensions[row_idx].height = 20

    # Tabela 2: Gradiente Térmico ISA T = 15 - 0.0065 * h
    ws_prec["A11"] = "2. Gradiente Térmico na Atmosfera Padrão (ISA): T = T₀ - (0,0065 * h)"
    ws_prec["A11"].font = font_section

    headers_isa = ["Nível de Voo / Altitude (m)", "Temperatura ao Nível do Mar T₀ (°C)", "Gradiente (°C/m)", "Fórmula no Excel", "Temperatura Estimada ISA (°C)"]
    for col_idx, text in enumerate(headers_isa, start=1):
        c = ws_prec.cell(row=12, column=col_idx, value=text)
        c.font = font_header; c.fill = fill_header_accent; c.alignment = align_header; c.border = border_cell
    ws_prec.row_dimensions[12].height = 25

    isa_dados = [
        ("0 m (Nível do Mar)", 0, 15.0, 0.0065),
        ("1.000 m (Tráfego)", 1000, 15.0, 0.0065),
        ("3.000 m (Navegação)", 3000, 15.0, 0.0065),
        ("11.000 m (Tropopausa)", 11000, 15.0, 0.0065),
    ]

    for row_idx, (desc, alt, t0, grad) in enumerate(isa_dados, start=13):
        zebra = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
        ws_prec.cell(row=row_idx, column=1, value=desc).font = font_bold; ws_prec.cell(row=row_idx, column=1).alignment = align_left; ws_prec.cell(row=row_idx, column=1).border = border_cell; ws_prec.cell(row=row_idx, column=1).fill = zebra
        c = ws_prec.cell(row=row_idx, column=2, value=t0); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.0 "°C"'
        c = ws_prec.cell(row=row_idx, column=3, value=grad); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.0000'
        c = ws_prec.cell(row=row_idx, column=4, value=f"=B{row_idx} - (C{row_idx} * {alt})"); c.font = font_formula; c.alignment = align_center; c.border = border_cell; c.fill = fill_formula_highlight
        c = ws_prec.cell(row=row_idx, column=5, value=f"=B{row_idx}-(C{row_idx}*{alt})"); c.font = font_bold; c.alignment = align_right; c.border = border_cell; c.fill = zebra; c.number_format = '0.0 "°C"'
        ws_prec.row_dimensions[row_idx].height = 20

    # Callout Armadilha da Potência e Negação
    ws_prec.merge_cells("A19:G23")
    ws_prec["A19"] = "⚠️ ATENÇÃO À ARMADILHA DO SINAL DE MENOS COM POTÊNCIA NO EXCEL:\n• No Excel, o operador unário - tem precedência sobre ^ : digitando =-5^2 o Excel calcula (-5)^2 = 25.\n• Na matemática padrão: -5^2 = -(5^2) = -25.\n• Boas Práticas: Use SEMPRE parênteses explícitos: =-(5^2) para -25 ou =(-5)^2 para 25!\n• Em fórmulas com subtração e multiplicação (como o ISA), o Excel multiplica antes de subtrair automaticamente, mas o uso de parênteses deixa a equação clara e segura."
    ws_prec["A19"].font = font_section; ws_prec["A19"].fill = fill_alert
    ws_prec["A19"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    autofit(ws_prec, 1, 5)

    # =========================================================================
    # ABA 6: ESTUDO DE CASO AERONÁUTICO — MÉDIA VS. MEDIANA E IMPACTO DE OUTLIERS
    # =========================================================================
    ws_stat = wb.create_sheet(title="Media_vs_Mediana_OS")
    ws_stat.views.sheetView[0].showGridLines = True

    ws_stat.merge_cells("A1:G1")
    ws_stat["A1"] = "ESTATÍSTICA DESCRITIVA NO HANGAR: DIFERENÇA ENTRE MÉDIA E MEDIANA DIANTE DE OUTLIERS"
    ws_stat["A1"].font = font_title; ws_stat["A1"].fill = fill_title; ws_stat["A1"].alignment = align_center
    ws_stat.row_dimensions[1].height = 30

    ws_stat["A3"] = "Relatório Semanal de Ordens de Serviço (OS) de Manutenção no Hangar"
    ws_stat["A3"].font = font_section

    headers_os = ["Nº da OS", "Aeronave", "Tipo de Manutenção", "Categoria", "Custo Total (R$)", "Tempo de Box (Horas)"]
    for col_idx, text in enumerate(headers_os, start=1):
        c = ws_stat.cell(row=4, column=col_idx, value=text)
        c.font = font_header; c.fill = fill_header; c.alignment = align_header; c.border = border_cell
    ws_stat.row_dimensions[4].height = 25

    os_dados = [
        ("OS-101", "Cessna 152 (PR-MNT)", "Troca de Óleo e Filtros", "Preventiva", 1200.00, 4),
        ("OS-102", "Cessna 172 (PR-ABC)", "Inspeção Periódica 50h", "Preventiva", 1500.00, 6),
        ("OS-103", "Piper Cherokee (PT-VLM)", "Revisão de Magnetos e Velas", "Preventiva", 1800.00, 8),
        ("OS-104", "Cirrus SR22 (PP-ZUL)", "Atualização e Calibração Aviônica", "Eletrônica", 2000.00, 5),
        ("OS-105", "Beech Baron (PR-FTE)", "Alinhamento e Balanceamento de Hélice", "Preventiva", 2500.00, 10),
        ("OS-106", "King Air C90 (PT-EXP)", "Overhaul Completo de Turbina PT6A (OUTLIER)", "Corretiva Pesada", 120000.00, 120),
    ]

    for row_idx, (num_os, aviao, servico, cat, custo, horas) in enumerate(os_dados, start=5):
        is_outlier = (row_idx == 10)
        row_fill = fill_outlier if is_outlier else (fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None))
        
        ws_stat.cell(row=row_idx, column=1, value=num_os).font = font_bold; ws_stat.cell(row=row_idx, column=1).alignment = align_center; ws_stat.cell(row=row_idx, column=1).border = border_cell; ws_stat.cell(row=row_idx, column=1).fill = row_fill
        ws_stat.cell(row=row_idx, column=2, value=aviao).font = font_data; ws_stat.cell(row=row_idx, column=2).alignment = align_left; ws_stat.cell(row=row_idx, column=2).border = border_cell; ws_stat.cell(row=row_idx, column=2).fill = row_fill
        ws_stat.cell(row=row_idx, column=3, value=servico).font = font_data; ws_stat.cell(row=row_idx, column=3).alignment = align_left; ws_stat.cell(row=row_idx, column=3).border = border_cell; ws_stat.cell(row=row_idx, column=3).fill = row_fill
        ws_stat.cell(row=row_idx, column=4, value=cat).font = font_data; ws_stat.cell(row=row_idx, column=4).alignment = align_center; ws_stat.cell(row=row_idx, column=4).border = border_cell; ws_stat.cell(row=row_idx, column=4).fill = row_fill
        
        c = ws_stat.cell(row=row_idx, column=5, value=custo); c.font = Font(name=font_family, size=10, bold=is_outlier, color="991B1B" if is_outlier else "1E293B"); c.alignment = align_right; c.border = border_cell; c.fill = row_fill; c.number_format = CURRENCY_FORMAT
        c = ws_stat.cell(row=row_idx, column=6, value=horas); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = row_fill; c.number_format = NUMBER_FORMAT
        ws_stat.row_dimensions[row_idx].height = 20

    # Linha Total
    ws_stat.merge_cells("A11:D11")
    ws_stat["A11"] = "TOTAL GERAL DAS ORDENS DE SERVIÇO"
    ws_stat["A11"].font = font_bold; ws_stat["A11"].alignment = align_right; ws_stat["A11"].fill = fill_total
    for col_c in range(1, 5):
        ws_stat.cell(row=11, column=col_c).border = border_cell; ws_stat.cell(row=11, column=col_c).fill = fill_total
    ws_stat["E11"] = "=SUM(E5:E10)"; ws_stat["E11"].font = font_bold; ws_stat["E11"].alignment = align_right; ws_stat["E11"].fill = fill_total; ws_stat["E11"].border = border_cell; ws_stat["E11"].number_format = CURRENCY_FORMAT
    ws_stat["F11"] = "=SUM(F5:F10)"; ws_stat["F11"].font = font_bold; ws_stat["F11"].alignment = align_center; ws_stat["F11"].fill = fill_total; ws_stat["F11"].border = border_cell; ws_stat["F11"].number_format = '#,##0 "h"'
    ws_stat.row_dimensions[11].height = 22

    # Linha Média
    ws_stat.merge_cells("A12:D12")
    ws_stat["A12"] = "MÉDIA ARITMÉTICA (=MÉDIA) — Sensível a Outliers"
    ws_stat["A12"].font = Font(name=font_family, size=10, bold=True, color="991B1B"); ws_stat["A12"].alignment = align_right; ws_stat["A12"].fill = fill_outlier
    for col_c in range(1, 5):
        ws_stat.cell(row=12, column=col_c).border = border_cell; ws_stat.cell(row=12, column=col_c).fill = fill_outlier
    ws_stat["E12"] = "=AVERAGE(E5:E10)"; ws_stat["E12"].font = Font(name=font_family, size=10, bold=True, color="991B1B"); ws_stat["E12"].alignment = align_right; ws_stat["E12"].fill = fill_outlier; ws_stat["E12"].border = border_cell; ws_stat["E12"].number_format = CURRENCY_FORMAT
    ws_stat["F12"] = "=AVERAGE(F5:F10)"; ws_stat["F12"].font = Font(name=font_family, size=10, bold=True, color="991B1B"); ws_stat["F12"].alignment = align_center; ws_stat["F12"].fill = fill_outlier; ws_stat["F12"].border = border_cell; ws_stat["F12"].number_format = '0.0 "h"'
    ws_stat.row_dimensions[12].height = 22

    # Linha Mediana
    ws_stat.merge_cells("A13:D13")
    ws_stat["A13"] = "MEDIANA (=MED) — Robusta / Representa o Serviço Típico"
    ws_stat["A13"].font = Font(name=font_family, size=10, bold=True, color="166534"); ws_stat["A13"].alignment = align_right; ws_stat["A13"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    for col_c in range(1, 5):
        ws_stat.cell(row=13, column=col_c).border = border_total; ws_stat.cell(row=13, column=col_c).fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    ws_stat["E13"] = "=MEDIAN(E5:E10)"; ws_stat["E13"].font = Font(name=font_family, size=10, bold=True, color="166534"); ws_stat["E13"].alignment = align_right; ws_stat["E13"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws_stat["E13"].border = border_total; ws_stat["E13"].number_format = CURRENCY_FORMAT
    ws_stat["F13"] = "=MEDIAN(F5:F10)"; ws_stat["F13"].font = Font(name=font_family, size=10, bold=True, color="166534"); ws_stat["F13"].alignment = align_center; ws_stat["F13"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws_stat["F13"].border = border_total; ws_stat["F13"].number_format = '0.0 "h"'
    ws_stat.row_dimensions[13].height = 22

    # Painel Comparativo Explicativo
    ws_stat.merge_cells("A15:F20")
    ws_stat["A15"] = "ANÁLISE GERENCIAL DA MANUTENÇÃO (MÉDIA vs. MEDIANA):\n\n1. MÉDIA (=MÉDIA(E5:E10) = R$ 21.500,00): A Média foi 'puxada' para cima pelo Overhaul da turbina (R$ 120.000). Nenhum cliente rotineiro pagou 21 mil! Use a média apenas para previsão de caixa financeiro total.\n\n2. MEDIANA (=MED(E5:E10) = R$ 1.900,00): A Mediana é o valor central do conjunto. Ela ignora os extremos e revela com extrema precisão o custo real de uma manutenção típica preventiva no hangar.\n\n3. REGRA DE ENGENHARIA: Use MEDIANA para precificar pacotes de serviços e estimar prazos de entrega padrão (Lead Time), evitando distorções causadas por eventos raros e pontuais."
    ws_stat["A15"].font = font_section; ws_stat["A15"].fill = fill_card
    ws_stat["A15"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    autofit(ws_stat, 1, 6)

    # =========================================================================
    # ABA 7: EXERCÍCIO DE FIXAÇÃO — CONSUMO DE TREINAMENTO
    # =========================================================================
    ws5 = wb.create_sheet(title="Exercicio_Fixacao_Consumo")
    ws5.views.sheetView[0].showGridLines = True

    ws5.merge_cells("A1:D1")
    ws5["A1"] = "PROGRAMA DE INSTRUÇÃO DE VOO — CONTROLE DE CONSUMO DE COMBUSTÍVEL"
    ws5["A1"].font = font_title; ws5["A1"].fill = fill_title; ws5["A1"].alignment = align_center
    ws5.row_dimensions[1].height = 30

    headers_treino = [
        "Piloto Aluno", "Horas de Voo (h)", "Consumo por Hora (L/h)", "Consumo Total de Combustível (L)"
    ]

    for col_idx, text in enumerate(headers_treino, start=1):
        cell = ws5.cell(row=3, column=col_idx, value=text)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_header; cell.border = border_cell
    ws5.row_dimensions[3].height = 28

    for row_idx, (aluno, horas, consumo_h) in enumerate(treinamento_dados, start=4):
        zebra = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
        c = ws5.cell(row=row_idx, column=1, value=aluno); c.font = font_data; c.alignment = align_left; c.border = border_cell; c.fill = zebra
        c = ws5.cell(row=row_idx, column=2, value=horas); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.0'
        c = ws5.cell(row=row_idx, column=3, value=consumo_h); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.0'
        c = ws5.cell(row=row_idx, column=4, value=f"=B{row_idx}*C{row_idx}"); c.font = font_bold; c.alignment = align_right; c.border = border_cell; c.fill = zebra; c.number_format = '#,##0.0 "L"'
        ws5.row_dimensions[row_idx].height = 20

    # Total Geral
    ws5["A10"] = "TOTAL DE COMBUSTÍVEL CONSUMIDO"; ws5["A10"].font = font_bold; ws5["A10"].alignment = align_right; ws5["A10"].fill = fill_total; ws5["A10"].border = border_cell
    ws5["B10"] = "=SUM(B4:B9)"; ws5["B10"].font = font_bold; ws5["B10"].alignment = align_center; ws5["B10"].fill = fill_total; ws5["B10"].border = border_cell; ws5["B10"].number_format = '0.0 "h"'
    ws5["C10"] = "-"; ws5["C10"].font = font_bold; ws5["C10"].alignment = align_center; ws5["C10"].fill = fill_total; ws5["C10"].border = border_cell
    ws5["D10"] = "=SUM(D4:D9)"; ws5["D10"].font = Font(name=font_family, size=10, bold=True, color=NAVY_HEX); ws5["D10"].alignment = align_right; ws5["D10"].fill = fill_total; ws5["D10"].border = border_cell; ws5["D10"].number_format = '#,##0.0 "L"'
    ws5.row_dimensions[10].height = 22

    # Média Geral
    ws5["A11"] = "MÉDIA POR ALUNO"; ws5["A11"].font = font_bold; ws5["A11"].alignment = align_right; ws5["A11"].fill = fill_card; ws5["A11"].border = border_cell
    ws5["B11"] = "=AVERAGE(B4:B9)"; ws5["B11"].font = font_bold; ws5["B11"].alignment = align_center; ws5["B11"].fill = fill_card; ws5["B11"].border = border_cell; ws5["B11"].number_format = '0.0 "h"'
    ws5["C11"] = "=AVERAGE(C4:C9)"; ws5["C11"].font = font_bold; ws5["C11"].alignment = align_center; ws5["C11"].fill = fill_card; ws5["C11"].border = border_cell; ws5["C11"].number_format = '0.0 "L/h"'
    ws5["D11"] = "=AVERAGE(D4:D9)"; ws5["D11"].font = Font(name=font_family, size=10, bold=True, color=NAVY_HEX); ws5["D11"].alignment = align_right; ws5["D11"].fill = fill_card; ws5["D11"].border = border_cell; ws5["D11"].number_format = '#,##0.0 "L"'
    ws5.row_dimensions[11].height = 22

    # Mediana Geral
    ws5["A12"] = "MEDIANA POR ALUNO"; ws5["A12"].font = font_bold; ws5["A12"].alignment = align_right; ws5["A12"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws5["A12"].border = border_total
    ws5["B12"] = "=MEDIAN(B4:B9)"; ws5["B12"].font = font_bold; ws5["B12"].alignment = align_center; ws5["B12"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws5["B12"].border = border_total; ws5["B12"].number_format = '0.0 "h"'
    ws5["C12"] = "=MEDIAN(C4:C9)"; ws5["C12"].font = font_bold; ws5["C12"].alignment = align_center; ws5["C12"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws5["C12"].border = border_total; ws5["C12"].number_format = '0.0 "L/h"'
    ws5["D12"] = "=MEDIAN(D4:D9)"; ws5["D12"].font = Font(name=font_family, size=10, bold=True, color="166534"); ws5["D12"].alignment = align_right; ws5["D12"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws5["D12"].border = border_total; ws5["D12"].number_format = '#,##0.0 "L"'
    ws5.row_dimensions[12].height = 22

    ws5.merge_cells("A14:D17")
    ws5["A14"] = "FÓRMULAS DO EXERCÍCIO DE FIXAÇÃO:\n• Consumo Total por Aluno (D4): =B4*C4 (Horas de Voo × Consumo Médio por Hora)\n• Total Geral de Horas e Litros (B10 e D10): =SOMA(B4:B9) e =SOMA(D4:D9)\n• Médias Operacionais (B11, C11 e D11): =MÉDIA(B4:B9), =MÉDIA(C4:C9) e =MÉDIA(D4:D9)\n• Medianas Centrais (B12, C12 e D12): =MED(B4:B9), =MED(C4:C9) e =MED(D4:D9)"
    ws5["A14"].font = font_section; ws5["A14"].fill = fill_card
    ws5["A14"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    autofit(ws5, 1, 4)

    # =========================================================================
    # ABA 8: GABARITO COMPLETO & PAINEL CONSOLIDADO
    # =========================================================================
    ws6 = wb.create_sheet(title="Gabarito_Completo")
    ws6.views.sheetView[0].showGridLines = True

    ws6.merge_cells("A1:I1")
    ws6["A1"] = "AEROCLUBE DE SOROCABA — PAINEL GERAL CONSOLIDADO (GABARITO OFICIAL)"
    ws6["A1"].font = font_title; ws6["A1"].fill = fill_title; ws6["A1"].alignment = align_center
    ws6.row_dimensions[1].height = 32

    # Tabela Hangar
    ws6["A3"] = "1. Controle Operacional de Hangar e Abastecimento"
    ws6["A3"].font = font_section

    for col_idx, text in enumerate(headers_hangar, start=1):
        cell = ws6.cell(row=4, column=col_idx, value=text)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_header; cell.border = border_cell
    ws6.row_dimensions[4].height = 28

    for row_idx, row_data in enumerate(frota_dados, start=5):
        zebra = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
        ws6.cell(row=row_idx, column=1, value=row_data[0]).alignment = align_center; ws6.cell(row=row_idx, column=1).font = font_bold
        ws6.cell(row=row_idx, column=2, value=row_data[1]).alignment = align_left; ws6.cell(row=row_idx, column=2).font = font_data
        c = ws6.cell(row=row_idx, column=3, value=row_data[2]); c.alignment = align_center; c.font = font_data; c.number_format = NUMBER_FORMAT
        c = ws6.cell(row=row_idx, column=4, value=row_data[3]); c.alignment = align_right; c.font = font_data; c.number_format = CURRENCY_FORMAT
        c = ws6.cell(row=row_idx, column=5, value=row_data[4]); c.alignment = align_center; c.font = font_data; c.number_format = NUMBER_FORMAT
        c = ws6.cell(row=row_idx, column=6, value=row_data[5]); c.alignment = align_right; c.font = font_data; c.number_format = CURRENCY_FORMAT
        cG = ws6.cell(row=row_idx, column=7, value=f"=C{row_idx}*D{row_idx}"); cG.alignment = align_right; cG.font = font_data; cG.number_format = CURRENCY_FORMAT
        cH = ws6.cell(row=row_idx, column=8, value=f"=E{row_idx}*F{row_idx}"); cH.alignment = align_right; cH.font = font_data; cH.number_format = CURRENCY_FORMAT
        cI = ws6.cell(row=row_idx, column=9, value=f"=G{row_idx}+H{row_idx}"); cI.alignment = align_right; cI.font = font_bold; cI.number_format = CURRENCY_FORMAT
        for col_c in range(1, 10):
            ws6.cell(row=row_idx, column=col_c).border = border_cell; ws6.cell(row=row_idx, column=col_c).fill = zebra
        ws6.row_dimensions[row_idx].height = 20

    ws6.merge_cells("A10:F10")
    ws6["A10"] = "TOTAL GERAL"; ws6["A10"].font = font_bold; ws6["A10"].alignment = align_right; ws6["A10"].fill = fill_total
    for col_c in range(1, 7):
        ws6.cell(row=10, column=col_c).border = border_total; ws6.cell(row=10, column=col_c).fill = fill_total
    ws6["G10"] = "=SUM(G5:G9)"; ws6["G10"].font = font_bold; ws6["G10"].alignment = align_right; ws6["G10"].border = border_total; ws6["G10"].fill = fill_total; ws6["G10"].number_format = CURRENCY_FORMAT
    ws6["H10"] = "=SUM(H5:H9)"; ws6["H10"].font = font_bold; ws6["H10"].alignment = align_right; ws6["H10"].border = border_total; ws6["H10"].fill = fill_total; ws6["H10"].number_format = CURRENCY_FORMAT
    ws6["I10"] = "=SUM(I5:I9)"; ws6["I10"].font = Font(name=font_family, size=10, bold=True, color=NAVY_HEX); ws6["I10"].alignment = align_right; ws6["I10"].border = border_total; ws6["I10"].fill = fill_total; ws6["I10"].number_format = CURRENCY_FORMAT
    ws6.row_dimensions[10].height = 22

    # Tabela Treino
    ws6["A13"] = "2. Controle de Consumo da Instrução Prática (Exercício de Fixação)"
    ws6["A13"].font = font_section

    for col_idx, text in enumerate(headers_treino, start=1):
        cell = ws6.cell(row=14, column=col_idx, value=text)
        cell.font = font_header; cell.fill = fill_header_accent; cell.alignment = align_header; cell.border = border_cell
    ws6.row_dimensions[14].height = 28

    for row_idx, (aluno, horas, consumo_h) in enumerate(treinamento_dados, start=15):
        zebra = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
        c = ws6.cell(row=row_idx, column=1, value=aluno); c.font = font_data; c.alignment = align_left; c.border = border_cell; c.fill = zebra
        c = ws6.cell(row=row_idx, column=2, value=horas); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.0'
        c = ws6.cell(row=row_idx, column=3, value=consumo_h); c.font = font_data; c.alignment = align_center; c.border = border_cell; c.fill = zebra; c.number_format = '0.0'
        c = ws6.cell(row=row_idx, column=4, value=f"=B{row_idx}*C{row_idx}"); c.font = font_bold; c.alignment = align_right; c.border = border_cell; c.fill = zebra; c.number_format = '#,##0.0 "L"'
        ws6.row_dimensions[row_idx].height = 20

    ws6["A21"] = "TOTAL DE COMBUSTÍVEL"; ws6["A21"].font = font_bold; ws6["A21"].alignment = align_right; ws6["A21"].fill = fill_total; ws6["A21"].border = border_cell
    ws6["B21"] = "=SUM(B15:B20)"; ws6["B21"].font = font_bold; ws6["B21"].alignment = align_center; ws6["B21"].fill = fill_total; ws6["B21"].border = border_cell; ws6["B21"].number_format = '0.0 "h"'
    ws6["C21"] = "-"; ws6["C21"].font = font_bold; ws6["C21"].alignment = align_center; ws6["C21"].fill = fill_total; ws6["C21"].border = border_cell
    ws6["D21"] = "=SUM(D15:D20)"; ws6["D21"].font = Font(name=font_family, size=10, bold=True, color=NAVY_HEX); ws6["D21"].alignment = align_right; ws6["D21"].fill = fill_total; ws6["D21"].border = border_cell; ws6["D21"].number_format = '#,##0.0 "L"'
    ws6.row_dimensions[21].height = 22

    ws6["A22"] = "MÉDIA POR ALUNO"; ws6["A22"].font = font_bold; ws6["A22"].alignment = align_right; ws6["A22"].fill = fill_card; ws6["A22"].border = border_cell
    ws6["B22"] = "=AVERAGE(B15:B20)"; ws6["B22"].font = font_bold; ws6["B22"].alignment = align_center; ws6["B22"].fill = fill_card; ws6["B22"].border = border_cell; ws6["B22"].number_format = '0.0 "h"'
    ws6["C22"] = "=AVERAGE(C15:C20)"; ws6["C22"].font = font_bold; ws6["C22"].alignment = align_center; ws6["C22"].fill = fill_card; ws6["C22"].border = border_cell; ws6["C22"].number_format = '0.0 "L/h"'
    ws6["D22"] = "=AVERAGE(D15:D20)"; ws6["D22"].font = Font(name=font_family, size=10, bold=True, color=NAVY_HEX); ws6["D22"].alignment = align_right; ws6["D22"].fill = fill_card; ws6["D22"].border = border_cell; ws6["D22"].number_format = '#,##0.0 "L"'
    ws6.row_dimensions[22].height = 22

    ws6["A23"] = "MEDIANA POR ALUNO"; ws6["A23"].font = font_bold; ws6["A23"].alignment = align_right; ws6["A23"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws6["A23"].border = border_total
    ws6["B23"] = "=MEDIAN(B15:B20)"; ws6["B23"].font = font_bold; ws6["B23"].alignment = align_center; ws6["B23"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws6["B23"].border = border_total; ws6["B23"].number_format = '0.0 "h"'
    ws6["C23"] = "=MEDIAN(C15:C20)"; ws6["C23"].font = font_bold; ws6["C23"].alignment = align_center; ws6["C23"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws6["C23"].border = border_total; ws6["C23"].number_format = '0.0 "L/h"'
    ws6["D23"] = "=MEDIAN(D15:D20)"; ws6["D23"].font = Font(name=font_family, size=10, bold=True, color="166534"); ws6["D23"].alignment = align_right; ws6["D23"].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"); ws6["D23"].border = border_total; ws6["D23"].number_format = '#,##0.0 "L"'
    ws6.row_dimensions[23].height = 22

    autofit(ws6, 1, 9)

    # Save Workbook
    out_path = r"c:\projetos\Material\Informática Aplicada a Aeronáutica\Exercicios\Aula_05_Excel_Basico_Hangar_e_Treinamento.xlsx"
    wb.save(out_path)
    print(f"Planilha atualizada com sucesso em: {out_path}")

if __name__ == "__main__":
    create_aula05_workbook()
